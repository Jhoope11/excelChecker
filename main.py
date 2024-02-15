from openpyxl import load_workbook
workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames


sheet = workbook.active
sheet


sheet.title


sheet["A1"]


sheet["A1"].value


sheet["F10"].value

sheet.cell(row=10, column=6)


sheet.cell(row=10, column=6).value

sheet["A1:C2"]
# Get all cells from column A
sheet["A"]






# Get all cells for a range of columns
sheet["A:B"]











# Get all cells from row 5
sheet[5]






# Get all cells for a range of rows
sheet[5:6]